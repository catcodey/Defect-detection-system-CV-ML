

#this one

import time
import os
import cv2
import shutil
from ultralytics import YOLO
import matplotlib.pyplot as plt
from pynput import keyboard
import csv 
from flask import Flask,render_template
import os
import threading
from pynput import keyboard

scanned_data = []  # To store the scanned data
def closecropped(results1,i,cropped_image,calc):
    for idx, result in enumerate(results1):
        box_count = 0  # To ensure you only process two boxes per image
        for box_id, box in enumerate(result.boxes):
            if box_count >= 2:  # Process only two bounding boxes
                break
            
            x1, y1, x2, y2 = box.xyxy[0].cpu().numpy()  # Get the coordinates
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)  # Convert to integer

            # Crop the image using the bounding box coordinates
            cropped_image1 = cropped_image[y1:y2, x1:x2]

            

            # Save the cropped image as a copy
            output_path = f'closecrop lock nut detetc_{idx}_{box_id}.jpg'
            cv2.imwrite(output_path, cropped_image1)
            
            print(f"Cropped up image saved as {output_path}")



            box_count += 1  # Increment the box counter


            # Pass the cropped image to the second model for further prediction
    if i==0:
        output_path="/Users/bbhavna/Desktop/CUMMINS PROJECT/closecrop lock nut detetc_0_0.jpg"
    else:
        output_path="/Users/bbhavna/Desktop/CUMMINS PROJECT/closecrop lock nut detetc_0_1.jpg"
    #save_pixels_to_csv(output_path, output_csv)#intensity excel sheet   #[FOR INTENSITY CHECKER ONLY] 
            
    #loosetightchker(calc)  #backup checks if its a tight or slightly loose screw  #[FOR INTENSITY CHECKER ONLY] 

    #locknutclassi(output_path)

    #template_matching(output_path)

    locknutdetect(output_path)
def locknutdetect(output_path):
    global lockcount
    from ultralytics import YOLO

    # Load the best classification model
    #model = YOLO(r"C:\Users\yd196\Downloads\best (21).pt")  # Adjust the path if needed
    model=YOLO(r"C:\Users\yd196\Desktop\lock nut final working\locknut closeup detetction best (31).pt")
    # Run inference on an image
    results = model(output_path,conf=0.7)

   

    class_names = model.names  # This should give you a list or dictionary of class names (index -> class)
    print("class names: ",class_names)
    # Define the class name or ID for the dataplate (you can also use the class index directly)
    locknut_class_name = "tight"  # Replace with your actual class name for dataplate

    locknut_found = False  # Flag to check if a dataplate is detected
    #lockcount=0
    # Iterate over the results to check for a dataplate
    for result in results:
        
        # `result.boxes` contains the bounding box information, including class IDs and confidence scores
        for box in result.boxes:
            print("BOX: ",box)
            class_id = int(box.cls[0])  # Get the class ID (index)
            print("CLASS ID: ",class_id)
            class_name = class_names[class_id]  # Convert the ID to class name
            
            # Check if the detected class is a dataplate
            if class_name == locknut_class_name:
                locknut_found = True
                print(f"locknut detected with confidence: {box.conf[0]}")  # Print confidence score too if needed
            
                lockcount+=1
            
            # If any dataplate was found, print it to the console

        
    

def save_pixels_to_csv(image_path, output_csv):
    
    # Load the image in grayscale
    image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    height, width = image.shape
    
    # Open the CSV file to write to
    with open(output_csv, 'w', newline='') as csvfile:
        csvwriter = csv.writer(csvfile)
        
        # Create a figure for displaying the rows
        #plt.figure(figsize=(10, 2))
        
        # Iterate over each row (scan line) in the image
        for y in range(height):
            # Extract the pixel values for the current scan line
            row_pixels = image[y, :]  # All pixel values in the y-th row
            
            # Convert the pixel values into a list and write to CSV
            csvwriter.writerow(row_pixels.tolist())

    
    print(f"Pixel values saved to {output_csv}")




def loosetightchker(calc):
    
    import pandas as pd
    global l
    cal=calc
    df=pd.read_csv(r"C:\Users\yd196\Desktop\intensity.csv")
    cal=(len(df.columns)//2)-20
    print(cal)
    #print(cal)
    column_values=df.iloc[0:40,cal-1]
    c=0
    t=0
    u=0
    chker=0
    for idx, i in enumerate(column_values):  # iterating/scanning the cal column vertically and getting the FIRST intensity value that is lesser than 40. 40 acts as threshold. all values below 40 leans towards black
        if i<=40:  #40 acts as threshold below which we encounter black pixels. i is first value below 40
            indexval=idx   #getting index of this i
            column_values1=df.iloc[indexval:indexval+3,cal-1]  #getting the next 5 rows of data  of the cal column. this is a sub dataframe
            chker=1
            break;
    if chker==0:
        l.append("tight as chker is 0")
        print("tight as chker is 0 ")
    else:
        print(column_values1)

        for k in range(cal-1,cal+16,1):  #horizontally iterating over 10 cells. this is the max no of cells we'll be checking this entire thing over
            if t==4:  # t indicates  no of sub dataframes in which all 5 rows have intensity value less than 40. it indicates  how many times/how long, along how many cols the intensity stays below 40
                break;
            else:
            
                for idx, i in column_values1.items():
                    print("idx value: ",idx)  #index value of each value of sub daatframe
                    u=u+1   # indicates no of times the for loop runs. its just for debugging purposes
                    if column_values1[idx]<=40:
                    
                        print("col idx: " ,column_values1[idx])
                    
                    
                        c=c+1   # this stores  the successive number of times the intensity is less than 40 i.e, each of the 5 values in the sub dataframe shd be less than 40 successsively
                        print("count value: ",c)
                    else: #ADDED [ENSURES ONLY SUCCESSIVE INTENSITY VALUES <40 ARE PICKED. OTHERWISE LOOP BROKEN]
                        c=0
                        break;
                print(u)
                if c>=3:  #if for one sub dataframe  all 5 successive values are less than 40, then c qill obvio be 5 and t is increamented. it means that this sub dataframe is considered
                    t=t+1
                print("t value: ",t)
                    
                column_values1=df.iloc[indexval:indexval+3, k+1]  # moving to the next sub dataframe horiznatlly. indices are the same 5 indices but we move one col to the right. this is done 10 times(cal-1 to cal+11).
                c=0
                print("hello")
                print(column_values1)
        print(t)
        if t==4:
            l.append("LOOSE")
            print("LOOSE")
        else:
            l.append("tight")
            print("tight")

def locknutclassi(output_path):
    global locknutclasscount
    # Load the best classification model
    model = YOLO(r"C:\Users\yd196\Downloads\best (21).pt")  # Adjust the path if needed

    # Run inference on an image
    results = model(output_path)

    # Access the prediction results
    pred = results[0]  # Access the first result (assuming only one image was processed)

    # Check if probabilities are available
    if pred.probs is not None:
        # Get the class probabilities as a numpy array
        probabilities = pred.probs.data.numpy()  # Convert to numpy array if it's a tensor

        # Get the index of the class with the highest probability
        top_class_index = probabilities.argmax()  # Index of the highest probability
        top_class_prob = probabilities[top_class_index]  # Probability of the highest class

        # Assuming the class mapping is as follows:
        class_names = pred.names  # Get the class names

        # Print the class with the highest probability
        print(f'Predicted class: {class_names[top_class_index]} with confidence: {top_class_prob:.2f}')
        if class_names[top_class_index]=="tight":
            #locknutclasscount+=1
            template_matching(output_path)
            #statuses.append(1)
       
    else:
        print("Probabilities are not available.")
        #statuses.append(1)
    

    im3="/Users/bbhavna/Desktop/CUMMINS PROJECT/runs/detect/classify"
    folder_to_delete = im3  # deleting predn folder

    # Check if the folder exists, then delete it
    if os.path.exists(folder_to_delete):
        shutil.rmtree(folder_to_delete)
        print(f"Folder '{folder_to_delete}' has been deleted.")
    else:
        print(f"Folder '{folder_to_delete}' does not exist.")

def template_matching(output_path):
    global tempmatchcount
    import cv2 as cv
    import numpy as np
    from matplotlib import pyplot as plt

    # Load the images in grayscale
    img = cv.imread(output_path, cv.IMREAD_GRAYSCALE)
    assert img is not None, "file could not be read, check with os.path.exists()"
    template = cv.imread(r'C:\Users\yd196\Desktop\template 6.jpg', cv.IMREAD_GRAYSCALE)
    assert template is not None, "file could not be read, check with os.path.exists()"

    # Get the dimensions of the image and template
    img_height, img_width = img.shape
    template_height, template_width = template.shape

    # Ensure that the template is smaller than the image in both dimensions
    if img_height < template_height or img_width < template_width:
        print("Template size is larger than the image. Resizing the template to fit the image.")
        # Resize the template to fit the image
        template = cv.resize(template, (img_width, img_height))
        template_height, template_width = template.shape

    w, h = template.shape[::-1]  # Get width and height of the (resized) template

    # Apply template matching to the full image
    method = cv.TM_CCOEFF  # Chosen template matching method
    res = cv.matchTemplate(img, template, method)
    min_val, max_val, min_loc, max_loc = cv.minMaxLoc(res)

    # Normalize the result
    res_norm = cv.normalize(res, None, 0, 1, cv.NORM_MINMAX, -1)

    # Set a threshold for detecting matches
    threshold = 0.95
    loc = np.where(res_norm >= threshold)
    
    # Draw rectangles on the full image for the matched regions
    matched_regions = []  # Store matched bounding box coordinates
    for pt in zip(*loc[::-1]):
        cv.rectangle(img, pt, (pt[0] + w, pt[1] + h), 255, 2)
        matched_regions.append((pt[0], pt[1], pt[0] + w, pt[1] + h))  # Bounding box (x1, y1, x2, y2)

    # Crop the image by half its height
    height, width = img.shape
    img_cropped = img[:height // 2, :]  # Top half of the image

    # Check if any bounding box is in the cropped image
    inside_cropped = False
    for (x1, y1, x2, y2) in matched_regions:
        if y1 < height // 2 and y2 < height // 2:
            inside_cropped = True
            print("At least one bounding box is inside the cropped region.")
            
            break  # Stop checking further if a bounding box is inside

    if not inside_cropped:
        tempmatchcount+=1
        print("No bounding boxes are inside the cropped region.")

    '''
    # Display the images using matplotlib
    plt.figure(figsize=(10, 5))
    
    # Plot full image with bounding boxes
    plt.subplot(1, 2, 1)
    plt.imshow(img, cmap='gray')
    plt.title('Full Image with Bounding Boxes')
    plt.axis('off')
    
    # Plot cropped image
    plt.subplot(1, 2, 2)
    plt.imshow(img_cropped, cmap='gray')
    plt.title('Cropped Image (Top Half)')
    plt.axis('off')
    
    plt.show()
    '''
    





        
   
        


def locknut():
    # STARTS HERE
    global lockcount
    for i in range(1):
        global l
        dictt={}
        st=time.time()
        lockcount=0
        print("LOCKCOUNT VALUE: ",lockcount)
        
        

        # Load the trained model

        # LOCK NUT WEIGHTS
        model1 = YOLO("/Users/bbhavna/Desktop/CUMMINS PROJECT/TRAINED_MODEL _16_12_24/TRAINED_MODEL/true weights for close up lock nut/locknut detetction yolov8s 100ep.pt")  # Replace with the path to your trained model weights
        #model1 = YOLO(r"C:\Users\yd196\Downloads\best (24).pt")
        model2=YOLO("TRAINED_MODEL _16_12_24/TRAINED_MODEL/true weights for close up lock nut/locknut close up 8n 100ep.pt")
        model3=YOLO("TRAINED_MODEL _16_12_24/TRAINED_MODEL/true weights for close up lock nut/locknut 8m 100ep best (21)(keypoint).pt")
        
        


        # Run prediction
        global tempmatchcount
        tempmatchcount=0
        global loosecount
        image = cv2.imread(image_path)
        results = model1.predict(source=image_path, save=False,conf=0.5)
       

        for idx, result in enumerate(results):
            for box_id, box in enumerate(result.boxes):

                loosecount=0
                x1, y1, x2, y2 = box.xyxy[0].cpu().numpy()  # Get the coordinates
                x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)  # Convert to integer

                # Crop the image using the bounding box coordinates
                cropped_image = image[y1:y2, x1:x2]
                
                # Save the cropped image as a copy
                output_path = f'cropped_image locknutidentification_{idx}_{box_id}.jpg'

                cv2.imwrite(output_path, cropped_image)
                
                print(f"Cropped image saved as {output_path}")
                
                results1 = model2.predict(source=cropped_image, save=True,conf=0.1)
                #ADDING THIS CROPPED IMG TO FLASK APP 
                #***************************************************************************
        
                
                im3 = '/Users/bbhavna/Desktop/CUMMINS PROJECT/runs/detect/predict'  # predn folder
                filesim1=os.listdir(im3)
                image_filename = filesim1[0]  # img name inside predn folder

               

                im1_copy_path = os.path.join(im3,  image_filename)    #  predn img path

                print(im1_copy_path)

                save_folder = '/Users/bbhavna/Desktop/CUMMINS PROJECT/prediction images folder'  # folder in which img to be saved
                
                new_image_name = "locknut.jpg"
                renamed_img_path = os.path.join(save_folder, new_image_name) 

                shutil.move(im1_copy_path, renamed_img_path)  #  move predn img in this new folder


                #renamed_img_path = os.path.join(save_folder, "nbbr.jpg")  # rename image.
                #os.rename(r"C:\Users\yd196\Desktop\FINAL PROJECT\prediction images folder\WIN_20240830_16_44_53_Pro.jpg", renamed_img_path)

                # Define the folder path you want to delete (e.g., 'runs/predict/exp')
                folder_to_delete = im3  # deleting predn folder

                # Check if the folder exists, then delete it
                if os.path.exists(folder_to_delete):
                    shutil.rmtree(folder_to_delete)
                    print(f"Folder '{folder_to_delete}' has been deleted.")
                else:
                    print(f"Folder '{folder_to_delete}' does not exist.")


                #******************************************************************


                 

                results2 = model3.predict(source=cropped_image, save=False,conf=0.2)
                

                for result in results2:
                    keypoints = result.keypoints  # Access keypoints from result
                    print(keypoints)
                    if keypoints is not None:
                        keypoints_array = keypoints.xy  # Assuming `.xy` contains the (x, y) coordinates

                        print(f"Keypoints for {image_path}:")
                        for i, (x, y) in enumerate(keypoints_array):
                            print(f"  Keypoint {i}: (x={x}, y={y})")
                            calc=(((x[0]-y[0])**2) + ((x[1]-y[1])**2))**0.5
                            dictt[i]=calc
                            print(calc)


                            if calc>20:
                                l.append("LOOSE as its >20")
                                print("LOOSE as its >20")
                                loosecount+=1
                                
                            
                            else:
                                
                                # Call the function with your image and output file path
                                # output_csv = r'C:\Users\yd196\Desktop\intensity.csv'  [FOR INTENSITY CHECKER ONLY]
                                
                                closecropped(results1,i,cropped_image,calc)

                '''
                if tempmatchcount>=2:  #and class lose >=2 
                    print("loose as tempmatchcount=2")
                    statuses.append(0)
                    #locknutcount=1
                elif tempmatchcount==1: #and class lose==1 #2 cases: 1 is looeecount is 1 for one keypoint and other keypoint tempmatchcount is 1
                    statuses.append(0)          # 2 is both keypoints went inside template matcher. but one of the points is loose in temp match checker.
                    #locknutcount=0
                elif  tempmatchcount==0: # 
                    statuses.append(1)
                                    
                '''
        if loosecount>1:
            statuses.append("NOT OK")
        elif lockcount>=2:
            statuses.append("OK")
        else:
            statuses.append("NOT OK")
 



    

def eclip():
    model4=YOLO("TRAINED_MODEL _16_12_24/TRAINED_MODEL/eclips/eclips 8s 100ep  best (24).pt/eclips 8s 100ep  best (24).pt") #E CLIP WEIGHTS 
    results3=model4.predict(source=image_path2,save=True,
                            conf=0.2)
    '''
    for re in results3:
        re.show()
    '''
    im1 = '/Users/bbhavna/Desktop/CUMMINS PROJECT/runs/detect/predict'  # predn folder
    filesim1=os.listdir(im1)
    image_filename = filesim1[0]  # img name inside predn folder

    

    im1_copy_path = os.path.join(im1,  image_filename)    #  predn img path

    print(im1_copy_path)

    save_folder = r'/Users/bbhavna/Desktop/CUMMINS PROJECT/prediction images folder'  # folder in which img to be saved
    
    new_image_name = "eclip.jpg"
    renamed_img_path = os.path.join(save_folder, new_image_name) 

    shutil.move(im1_copy_path, renamed_img_path)  #  move predn img in this new folder


    
    # Define the folder path you want to delete (e.g., 'runs/predict/exp')
    folder_to_delete = im1  # deleting predn folder

    # Check if the folder exists, then delete it
    if os.path.exists(folder_to_delete):
        shutil.rmtree(folder_to_delete)
        print(f"Folder '{folder_to_delete}' has been deleted.")
    else:
        print(f"Folder '{folder_to_delete}' does not exist.")
    

    class_names = model4.names  # This should give you a list or dictionary of class names (index -> class)
    print("class names: ",class_names)
    # Define the class name or ID for the dataplate (you can also use the class index directly)
    eclip_class_name = "eclip-yes"  # Replace with your actual class name for dataplate

    eclip_found = False  # Flag to check if a dataplate is detected
    eclipcount=0
    # Iterate over the results to check for a dataplate
    for result in results3:
        
        # `result.boxes` contains the bounding box information, including class IDs and confidence scores
        for box in result.boxes:
            print("BOX: ",box)
            class_id = int(box.cls[0])  # Get the class ID (index)
            print("CLASS ID: ",class_id)
            class_name = class_names[class_id]  # Convert the ID to class name
            
            # Check if the detected class is a dataplate
            if class_name == eclip_class_name:
                eclip_found = True
                print(f"eclip detected with confidence: {box.conf[0]}")  # Print confidence score too if needed
            
                eclipcount+=1
            # If any dataplate was found, print it to the console
    if eclipcount>=1:
        
        print("eclip present.")
        statuses.append("OK")
    else:
        print("eclip not found.")
        statuses.append("NOT OK")
    print("status update for eclip: ",statuses)



    



def contrast_adjusted(saved_image_path):
    import numpy as np
    import matplotlib.pyplot as plt
    from skimage import io, exposure, color

    # Load the image and convert it to grayscale
    original = io.imread(saved_image_path)
    gray_image = color.rgb2gray(original)  # Convert to grayscale

    # Stretch contrast
    # Normalize the image to [0, 1] range
    normalized_image = exposure.rescale_intensity(gray_image, in_range='image')

    # Enhance contrast by adjusting the gamma
    contrast_adjusted = exposure.adjust_gamma(normalized_image, gamma=5)  # Lower gamma increases brightness

    # Scale the image back to [0, 255] for saving
    contrast_adjusted_255 = (contrast_adjusted * 255).astype(np.uint8)

    # Save the image

    import cv2
    output_path2 = f'contrast_adjusted_image.jpg'
    cv2.imwrite(output_path2, contrast_adjusted_255)

    #cv2.imwrite(r"C:\Users\yd196\Desktop\output1.jpg", contrast_adjusted_255)

    # Display the results
    
    '''
    plt.subplot(1, 2, 1)
    plt.title('Original Grayscale Image')
    plt.imshow(gray_image, cmap='gray')
    plt.axis('off')

    plt.subplot(1, 2, 2)
    plt.title('Contrast Adjusted Image')
    plt.imshow(contrast_adjusted, cmap='gray')
    plt.axis('off')

    plt.show()
    '''
    canny(output_path2)

def canny(contrast_adj_img):
    import numpy as np
    import cv2 as cv
    import matplotlib.pyplot as plt

    # Load the image in grayscale
    img = cv.imread(contrast_adj_img, cv.IMREAD_GRAYSCALE)
    assert img is not None, "file could not be read, check with os.path.exists()"

    # Resize the image (optional)
    img = cv.resize(img, (320, 320))


    height,widtth = img.shape
    img = img[:(height // 2)-25, :(widtth//2)+120]

    # Apply Gaussian blur to reduce noise and improve edge detection
    blurred = cv.GaussianBlur(img, (5, 5), 0)

    # Apply Canny edge detection
    edges = cv.Canny(blurred, 100, 300)# if u reduce second value you'll get more details

    # Display the original image and the edges
    #plt.figure(figsize=(6, 6))

    # Original image
    '''
    plt.subplot(1, 3, 1)
    plt.imshow(img, cmap='gray')
    plt.title('Original Image')
    plt.axis('off')

    # Canny edges
    plt.subplot(1, 3, 2)
    plt.imshow(edges, cmap='gray')
    plt.title('Canny Edges')
    plt.axis('off')
    
    '''


    edges_present = np.any(edges > 0)
    
    if edges_present==True:
        statuses.append("OK")
    elif edges_present==False:
        statuses.append("NOT OK")
    print("sttaus update for elc: ",statuses)

    '''
    # Display whether edges are present in the ROI
    plt.subplot(1, 3, 3)
    plt.text(0.5, 0.5, f'Edges Present: {edges_present}', fontsize=15, ha='center', va='center')
    plt.axis('off')
    plt.show()
    '''
    print("status update for elc: ",statuses)
    
def crop_elc():
    image2 = cv2.imread(image_path2)
    model5=YOLO("TRAINED_MODEL _16_12_24/TRAINED_MODEL/elc/elc yolov8n 100ep best (21).pt")
    results5=model5.predict(source=image_path2,save=False,conf=0.5)
 
    for idx, result in enumerate(results5):
        for box_id, box in enumerate(result.boxes):
            x1, y1, x2, y2 = box.xyxy[0].cpu().numpy()  # Get the coordinates
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)  # Convert to integer

            # Crop the image using the bounding box coordinates
            cropped_image = image2[y1:y2, x1:x2]
            
            # Save the cropped image as a copy
            output_path = f'cropped_image elc identification_{idx}_{box_id}.jpg'
            output_path1 = f'cropped_image elc identificationnn_{idx}_{box_id}.jpg'
            output_path1_cropped=f'cropped_cropped_image elc identificationnn_{idx}_{box_id}.jpg'
            cropped_croppedimage=cropped_image[:cropped_image.shape[0] // 2, :]
            cv2.imwrite(output_path, cropped_image)
            cv2.imwrite(output_path1, cropped_image)
            cv2.imwrite(output_path1_cropped, cropped_croppedimage)
            print(f"Cropped image saved as {output_path1}")

            
            save_folder = '/Users/bbhavna/Desktop/CUMMINS PROJECT/prediction images folder'  # folder in which img to be saved
            
            new_image_name = "elc.jpg"
            renamed_img_path = os.path.join(save_folder, new_image_name) 

            shutil.move(output_path1_cropped, renamed_img_path)  #  move predn(output_path) img in this new folder(renamed_img_path)
            

    contrast_adjusted(output_path1)
    



def elc():
    crop_elc()
    
def drivescrew(image_path1):
    model6 = YOLO("TRAINED_MODEL _16_12_24/TRAINED_MODEL/drive screw+daatplate/best (27).pt")  # Replace with the path to your trained model weights
    model7=YOLO("TRAINED_MODEL _16_12_24/TRAINED_MODEL/drive screw+daatplate/dataplate drivescrew yolov8m 100ep best (27).pt")
    image3 = cv2.imread(image_path1)
    results6 = model6.predict(source=image_path1, save=False,conf=0.2)
    for idx, result in enumerate(results6):
        for box_id, box in enumerate(result.boxes):
            x1, y1, x2, y2 = box.xyxy[0].cpu().numpy()  # Get the coordinates
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)  # Convert to integer

            # Crop the image using the bounding box coordinates
            cropped_image = image3[y1:y2, x1:x2]
            
            # Save the cropped image as a copy
            output_path = f'cropped_image for drive screw_{idx}_{box_id}.jpg'
            cv2.imwrite(output_path, cropped_image)
            
            print(f"Cropped image saved as {output_path}")
            

    results7 = model7.predict(source=cropped_image, save=True,conf=0.1)

        # Get the class names used in the model
    class_names = model7.names  # This should give you a list or dictionary of class names (index -> class)
    print("class names: ",class_names)
    # Define the class name or ID for the dataplate (you can also use the class index directly)
    drivescrew_class_name = "drive-screw-yes"  # Replace with your actual class name for dataplate

    drivescrew_found = False  # Flag to check if a dataplate is detected
    drivecount=0
    # Iterate over the results to check for a dataplate
    for result in results7:
        
        # `result.boxes` contains the bounding box information, including class IDs and confidence scores
        for box in result.boxes:
            print("BOX: ",box)
            class_id = int(box.cls[0])  # Get the class ID (index)
            print("CLASS ID: ",class_id)
            class_name = class_names[class_id]  # Convert the ID to class name
            
            # Check if the detected class is a dataplate
            if class_name == drivescrew_class_name:
                drivescrew_found = True
                print(f"Drivescrew detected with confidence: {box.conf[0]}")  # Print confidence score too if needed
            
                drivecount+=1
            # If any dataplate was found, print it to the console
    if drivecount>=2:
        
        print("Drivescrew present.")
        statuses.append("OK")
    else:
        print("Drivescrew not found.")
        statuses.append("NOT OK")
    print("status update for drive: ",statuses)
    
    #FOR FLASK APP
    #*****************************************************************

    im4 = '/Users/bbhavna/Desktop/CUMMINS PROJECT/runs/detect/predict'  # predn folder
    filesim1=os.listdir(im4)
    image_filename = filesim1[0]  # img name inside predn folder

    

    im1_copy_path = os.path.join(im4,  image_filename)    #  predn img path

    print(im1_copy_path)

    save_folder = '/Users/bbhavna/Desktop/CUMMINS PROJECT/prediction images folder'  # folder in which img to be saved
    
    new_image_name = "drivescrew.jpg"
    renamed_img_path = os.path.join(save_folder, new_image_name) 

    shutil.move(im1_copy_path, renamed_img_path)  #  move predn img in this new folder


    
    folder_to_delete = im4  # deleting predn folder

    # Check if the folder exists, then delete it
    if os.path.exists(folder_to_delete):
        shutil.rmtree(folder_to_delete)
        print(f"Folder '{folder_to_delete}' has been deleted.")
    else:
        print(f"Folder '{folder_to_delete}' does not exist.")


    #*******************************************************************
    '''
    for result in results7:
        result.show()
    '''

def nbbr(image_path1):
    model8=YOLO("TRAINED_MODEL _16_12_24/TRAINED_MODEL/nbbr/best.pt nbbr yolov8s 100ep.pt")
    result9=model8.predict(source=image_path1, save=True,conf=0.7)
    '''
    for r in result9:
        r.show()
    '''
    #FOR FLASK APP ONLY
    #*************************************************************
    im1 = '/Users/bbhavna/Desktop/CUMMINS PROJECT/runs/detect/predict'  # predn folder
    filesim1=os.listdir(im1)
    image_filename = filesim1[0]  # img name inside predn folder

    

    im1_copy_path = os.path.join(im1,  image_filename)    #  predn img path

    print(im1_copy_path)

    save_folder = '/Users/bbhavna/Desktop/CUMMINS PROJECT/prediction images folder'  # folder in which img to be saved
    
    new_image_name = "nbbr.jpg"
    renamed_img_path = os.path.join(save_folder, new_image_name) 

    shutil.move(im1_copy_path, renamed_img_path)  #  move predn img in this new path


    
    folder_to_delete = im1  # deleting predn folder

    # Check if the folder exists, then delete it
    if os.path.exists(folder_to_delete):
        shutil.rmtree(folder_to_delete)
        print(f"Folder '{folder_to_delete}' has been deleted.")
    else:
        print(f"Folder '{folder_to_delete}' does not exist.")

    class_names = model8.names  # This should give you a list or dictionary of class names (index -> class)
    print("class names: ",class_names)
    # Define the class name or ID for the dataplate (you can also use the class index directly)
    nbbr_class_name = "nbrr-present"  # Replace with your actual class name for dataplate

    nbbr_found = False  # Flag to check if a dataplate is detected
    nbbrcount=0
    # Iterate over the results to check for a dataplate
    for result in result9:
        
        # `result.boxes` contains the bounding box information, including class IDs and confidence scores
        for box in result.boxes:
            print("BOX: ",box)
            class_id = int(box.cls[0])  # Get the class ID (index)
            print("CLASS ID: ",class_id)
            class_name = class_names[class_id]  # Convert the ID to class name
            
            # Check if the detected class is a dataplate
            if class_name == nbbr_class_name:
                nbbr_found = True
                print(f"nbbr detected with confidence: {box.conf[0]}")  # Print confidence score too if needed
            
                nbbrcount+=1
            
            # If any dataplate was found, print it to the console
    if nbbrcount>=1:
        
        print("nbbr present.")
        statuses.append("OK")
        statuses.append("OK")
    else:
        print("nbbr not found.")
        statuses.append("NOT OK")
        statuses.append("NOT OK")
    print("status update for nbbr: ",statuses)
    #**************************************************************************
def cameracapture(save_path):
    
    global image_path1,image_path2,image_path
    import cv2
    import time
    import os

    # Global variables
   
    image_path1 = os.path.join(save_path, "camera1_image.jpg")

    # Create the directory if it doesn't exist
    if not os.path.exists(save_path):
        os.makedirs(save_path)

    # Open video capture for the webcam (0 for default camera)
    cap1 = cv2.VideoCapture(0)

    # Check if the camera opened successfully
    if not cap1.isOpened():
        print("Error: Could not open the camera.")
        exit()

    start_time = time.time()

    # Disable auto-focus and manually set focus for the webcam
    cap1.set(cv2.CAP_PROP_AUTOFOCUS, 0)  # 0 means manual focus

    # Set focus and resolution for the webcam
    cap1.set(cv2.CAP_PROP_FOCUS, 100)  # Adjust the focus value
    cap1.set(3, 1280)  # Set width

    capture_after_seconds = 20  # Time delay for capturing image

    while True:
        # Read frame from the webcam
        ret1, frame1 = cap1.read()

        if not ret1:
            print("Error: Could not read frame from the camera.")
            break

        # Calculate elapsed time
        elapsed_time = time.time() - start_time

        # Capture image after the specified time
        if elapsed_time >= capture_after_seconds:
            # Use global variable for the image path
            cv2.imwrite(image_path1, frame1)  # Save frame from the camera

            print(f"Image from the camera saved at {image_path1}")
            break  # Exit the loop after capturing the image

        # Press 'q' to exit
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release the video capture object and close windows
    cap1.release()
    cv2.destroyAllWindows()



    

    '''
    global image_path1,image_path2,image_path
    cap1 = cv2.VideoCapture(0)  # First camera
    cap2 = cv2.VideoCapture(1)  # Second camera

    # Check if both cameras opened successfully
    if not cap1.isOpened() or not cap2.isOpened():
        print("Error: Could not open one or both cameras.")
        return

    # Capture one frame from each camera
    ret1, frame1 = cap1.read()
    ret2, frame2 = cap2.read()

    if ret1 and ret2:
        # Ensure the directory exists
        if not os.path.exists(save_path):
            os.makedirs(save_path)

        # Save the captured frames as images in the same folder with different names
        image_path1 = os.path.join(save_path, "camera1_image.jpg")
        image_path2 = os.path.join(save_path, "camera2_image.jpg")

        cv2.imwrite(image_path1, frame1)
        cv2.imwrite(image_path2, frame2)

        print(f"Image from camera 1 saved at {image_path1}")
        print(f"Image from camera 2 saved at {image_path2}")
    else:
        print("Error: Could not capture images from one or both cameras.")

    # Release the cameras
    cap1.release()
    cap2.release()
    '''


import os
import threading
from pynput import keyboard
import shutil
import time
from flask import Flask, render_template

scanned_data = []  # List to hold scanned data
file_path = "/Users/bbhavna/Desktop/CUMMINS PROJECT/scanned_text_1.txt"



def main_function_2():
    global image_path, image_path1, image_path2, c2,statuses,new_data,image_path2_copy
    #time.sleep(20)  # Simulating a process
    st2 = time.time()
    new_data=[] #excel list containing all values of each defect
    # Example image paths
    image_path = "images used/camera42_image.jpg"
    image_path1 = "images used/camera41_image.jpg"
    image_path2 = "images used/WIN_20241019_16_16_19_Pro.jpg"
    image_path2_copy="images used/WIN_20241019_16_16_19_Pro.jpg"

    save_folder = "/Users/bbhavna/Desktop/CUMMINS PROJECT/captured images new"

    # Capture and save images from both cameras in the same folder
    #cameracapture(save_folder)
    
    # Call your model-related functions here
    
    
    
    drivescrew(image_path1)
    #locknut()
    elc()
    nbbr(image_path1)
    print("STATUSES: ",statuses)
    from datetime import datetime

    # Get the current date and time
    current_datetime = datetime.now()

    # Format the date and time numerically and round seconds
    formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
    print(formatted_datetime)
    new_data.append([partno]+statuses+[formatted_datetime])
    print("NEW DATA: ",new_data)
    excelread_2()
    
    if "NOT OK" in statuses:
        print("DEFECT")
       
        bin_output=r"C:\Users\yd196\Desktop\FINAL PROJECT\my flask folder\status_output_file.txt"
        with open(bin_output, "w") as bo:
            bo.write('1')
        
        #import winsound
        #alarm_sound = r'C:\Users\yd196\Downloads\slot-machine-payout-alarm-1996.wav'
        #winsound.PlaySound(alarm_sound, winsound.SND_FILENAME | winsound.SND_LOOP)
        #time.sleep(0.5)  # Let the sound play for 5 seconds
        #winsound.PlaySound(None, winsound.SND_PURGE)  # Stop the sound
        
    else:
        print("NO DEFECT")
        bin_output=r"C:\Users\yd196\Desktop\FINAL PROJECT\my flask folder\status_output_file.txt"
        with open(bin_output, "w") as bo:
            bo.write('0')
        

    time.sleep(10)
    statuses=[] #statuses list cleared   #UN COMMENT THIS LATER
    new_data=[]   #UN COMMENT THIS LATER

    # Copy files after model prediction
    source_folder = r'C:\Users\yd196\Desktop\FINAL PROJECT\prediction images folder'
    destination_folder = r'C:\Users\yd196\Desktop\FINAL PROJECT\my flask folder\static\images'

    for filename in os.listdir(source_folder):
        source_image_path = os.path.join(source_folder, filename)
        if os.path.isfile(source_image_path) and filename.lower().endswith(('.png', '.jpg', '.jpeg')):
            destination_image_path = os.path.join(destination_folder, filename)
            shutil.copy(source_image_path, destination_image_path)
            print(f"Copied {filename} to {destination_folder}")

    print("All images copied successfully.")
    et2 = time.time()
    print("ESTIMATED OVERALL TIME: ", et2 - st2)

    c2 = 0
    print("Resetting c2 to 0")

def main_function():
    global image_path, image_path1, image_path2, c1,statuses,new_data,image_path2_copy
    #time.sleep(20)  # Simulating a process
    st2 = time.time()
    new_data=[] #excel list containing all values of each defect
    # Example image paths
    image_path = "images used/camera42_image.jpg"
    image_path1 = "images used/camera41_image.jpg"
    image_path2 = "images used/WIN_20241019_16_16_19_Pro.jpg"
    image_path2_copy="images used/WIN_20241019_16_16_19_Pro.jpg"

    save_folder = "/Users/bbhavna/Desktop/CUMMINS PROJECT/captured images new"

    # Capture and save images from both cameras in the same folder
    #cameracapture(save_folder)

    # Call your model-related functions here
    
    
    eclip()
    drivescrew(image_path1)
    locknut()
    elc()
    nbbr(image_path1)
    print("STATUSES: ",statuses)
    from datetime import datetime

    # Get the current date and time
    current_datetime = datetime.now()

    # Format the date and time numerically and round seconds
    formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")
    print(formatted_datetime)
    new_data.append([partno]+statuses+[formatted_datetime])
    print("NEW DATA: ",new_data)
    excelread()
    
    if "NOT OK" in statuses:
        print("DEFECT")
       
        bin_output=r"my flask folder/status_output_file.txt"
        with open(bin_output, "w") as bo:
            bo.write('1')
        
        #import winsound
        #alarm_sound = r'C:\Users\yd196\Downloads\slot-machine-payout-alarm-1996.wav'
        #winsound.PlaySound(alarm_sound, winsound.SND_FILENAME | winsound.SND_LOOP)
        #time.sleep(0.5)  # Let the sound play for 5 seconds
        #winsound.PlaySound(None, winsound.SND_PURGE)  # Stop the sound
        
    else:
        print("NO DEFECT")
        bin_output=r"my flask folder/status_output_file.txt"
        with open(bin_output, "w") as bo:
            bo.write('0')
        

    time.sleep(10)
    statuses=[] #statuses list cleared   #UN COMMENT THIS LATER
    new_data=[]   #UN COMMENT THIS LATER

    # Copy files after model prediction
    source_folder = '/Users/bbhavna/Desktop/CUMMINS PROJECT/prediction images folder'
    destination_folder = 'my flask folder/static/images'

    for filename in os.listdir(source_folder):
        source_image_path = os.path.join(source_folder, filename)
        if os.path.isfile(source_image_path) and filename.lower().endswith(('.png', '.jpg', '.jpeg')):
            destination_image_path = os.path.join(destination_folder, filename)
            shutil.copy(source_image_path, destination_image_path)
            print(f"Copied {filename} to {destination_folder}")

    print("All images copied successfully.")
    et2 = time.time()
    print("ESTIMATED OVERALL TIME: ", et2 - st2)

    c1 = 0
    print("Resetting c1 to 0")
    
def excelread():
    global new_data,partno
    from openpyxl import load_workbook
    import os

    # Define the file path
    file_path = "my flask folder/trialbook1.xlsx"

    # Data to add: tuple format (nbbr, eclip, locknut, elc, drivescrew)
    #new_data = []
    #new_data.append([partno]+statuses)
    # Check if the Excel file exists
    if os.path.exists(file_path):
        # Load the existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        raise FileNotFoundError(f"The file {file_path} does not exist. Please create it with headers.")

    # Find the next empty row after the headers
    next_row = sheet.max_row + 1

    # Append the new 0/1 values under the corresponding columns (without touching headers)
    for row in new_data:
        sheet.append(row)

    # Save the workbook after appending the new rows
    workbook.save(file_path)

    print(f"0/1 values have been added to {file_path} starting from row {next_row}.")

def excelread_2():
    global new_data,partno
    from openpyxl import load_workbook
    import os

    # Define the file path
    file_path = r"C:\Users\yd196\Desktop\FINAL PROJECT\my flask folder\trialbook2.xlsx"

    # Data to add: tuple format (nbbr, eclip, locknut, elc, drivescrew)
    #new_data = []
    #new_data.append([partno]+statuses)
    # Check if the Excel file exists
    if os.path.exists(file_path):
        # Load the existing workbook
        workbook = load_workbook(file_path)
        sheet = workbook.active
    else:
        raise FileNotFoundError(f"The file {file_path} does not exist. Please create it with headers.")

    # Find the next empty row after the headers
    next_row = sheet.max_row + 1

    # Append the new 0/1 values under the corresponding columns (without touching headers)
    for row in new_data:
        sheet.append(row)

    # Save the workbook after appending the new rows
    workbook.save(file_path)

    print(f"0/1 values have been added to {file_path} starting from row {next_row}.")
def is_file_empty(file_path):
    """Check if the file is empty."""
    return os.stat(file_path).st_size == 0

def on_press(key):
    """Handle keyboard events."""
    global scanned_data
    print("Inside on_press")

    if hasattr(key, 'char') and key.char is not None:
        scanned_data.append(key.char)
        print("Character added to list:", scanned_data)
    
    elif key == keyboard.Key.enter:
        data = ''.join(scanned_data)
        with open(file_path, 'a') as f:
            print("Writing data to file:", scanned_data)
            f.write(data + '\n')

        print("Data written to file:", scanned_data)
        scanned_data.clear()
        print("Data cleared after saving")

def keyboard_listener():
    """Run the keyboard listener in a separate thread."""
    with keyboard.Listener(on_press=on_press) as listener:
        listener.join()

def monitor_file_and_run_model():
    global partno,user_choice
    """Continuously check if file has contents and run model if it does."""
    while True:
        if not is_file_empty(file_path):
            print("File has contents, running the main function...")
            #time.sleep(10)
            # Clear the file after running the main function
            # Open the file and read the line
            with open('scanned_text_1.txt', 'r') as file:
                # Read the first line from the file
                line = file.readline().strip()

                # Split the line using commas as the delimiter
                values = line.split(',')

                # Print the list of values (they will be strings automatically)
                partno=values[2]

            with open(file_path, "w") as a:
                a.write('')

            bin_output="my flask folder/status_output_file.txt"
            with open(bin_output, "w") as bo:
                bo.write(' ')
            #time.sleep(10)
            # Start the main function in a separate thread
            if user_choice=="144":
                print("1444 selected")
                # Start the Flask application in a separate thread
                flask_thread = threading.Thread(target=run_flask_144, daemon=True)
                flask_thread.start()
                main_thread = threading.Thread(target=main_function_2)
                main_thread.start()
            else:
                flask_thread = threading.Thread(target=run_flask_normal, daemon=True)
                flask_thread.start()
                main_thread = threading.Thread(target=main_function)
                main_thread.start()

            
# Function to run the 144 Flask app
def run_flask_144():
    app_144.run(debug=True, use_reloader=False)
# Flask app for "144" functionality
app_144 = Flask(__name__)

@app_144.route('/')
def home_144():
    print("144444444444")
    import pandas as pd
    global statuses
    # Example image filenames
    image_filenames = [ 
        'drivescrew.jpg',
        'locknut.jpg',
        'elc.jpg',
        'nbbr.jpg'
        
    ]
    
    # Example statuses (1 for OK, 0 for NOT OK)
    #statuses = [] # Example binary values for each image

    # Zip images and statuses together into pairs

    
    images_with_status = zip(image_filenames, statuses)
    excel_file = 'my flask folder/trialbook2.xlsx'
   
    # Read the Excel file into a DataFrame
    df = pd.read_excel(excel_file)



    # Sort the DataFrame in descending order based on the 'time' column
    df = df.sort_values(by='datetime', ascending=False)

    
    # Convert the DataFrame to HTML
    excel_data = df.to_html(classes='table table-striped', index=False)

    return render_template('index_2.html',data=excel_data, images_with_status=images_with_status)

   
# Flask app for "Normal" functionality
app_normal = Flask(__name__)

@app_normal.route('/')
def home_normal():
    import pandas as pd
    global statuses
    # Example image filenames
    image_filenames = [ 
        'eclip.jpg', 
        'drivescrew.jpg',
        'locknut.jpg',
        'elc.jpg',
        'nbbr.jpg'
        
    ]
    
    # Example statuses (1 for OK, 0 for NOT OK)
    #statuses = [] # Example binary values for each image

    # Zip images and statuses together into pairs

    
    images_with_status = zip(image_filenames, statuses)
    excel_file = r'my flask folder/trialbook1.xlsx'
   
    # Read the Excel file into a DataFrame
    df = pd.read_excel(excel_file)



    # Sort the DataFrame in descending order based on the 'time' column
    df = df.sort_values(by='datetime', ascending=False)

    
    # Convert the DataFrame to HTML
    excel_data = df.to_html(classes='table table-striped', index=False)

    return render_template('index_2.html',data=excel_data, images_with_status=images_with_status)




# Function to run the Normal Flask app
def run_flask_normal():
    app_normal.run(debug=True, use_reloader=False)



def set_choice(choice):
    """Set the user's choice and print it."""
    global user_choice
    user_choice = choice
    print(f"User choice set to: {user_choice}")

def create_button_interface():
    from tkinter import Tk,Button
    """Create a full-screen GUI with large buttons for user input."""
    global root
    root = Tk()
    root.title("Select Mode")
    root.geometry("1000x1000") 

    # Button for "144"
    button_144 = Button(
        root, text="144", command=lambda: set_choice("144"),
        font=("Arial", 48), width=10, height=2
    )
    button_144.pack(pady=100)

    # Button for "Normal"
    button_normal = Button(
        root, text="Normal", command=lambda: set_choice("Normal"),
        font=("Arial", 48), width=10, height=2
    )
    button_normal.pack(pady=100)
    
    
    root.mainloop()

def main():
    """Main function to run the listener,,j
      file monitor, and Flask app in parallel."""
    # Start the keyboard listener in a separate thread
    listener_thread = threading.Thread(target=keyboard_listener, daemon=True)
    listener_thread.start()



    button_thread = threading.Thread(target=create_button_interface, daemon=True)
    button_thread.start()
    # Start the file monitor and model runner in the main thread
    monitor_file_and_run_model()



    




        #else:
        #    print("File is empty, waiting for contents...")
        
        #time.sleep(1)  # Check every 1 second



def MAIN_FUNCTION():   #MAIN FN
    global scanned_data,l,output_csv,statuses,user_choice
    user_choice = None
    scanned_data=[]
    statuses=[]
    l=[]
    output_csv = r'C:\Users\yd196\Desktop\intensity.csv'
    
    '''
    print("initialising camefa...")
    st=time.time()
    cap=cv2.VideoCapture(0)
    et=time.time()
    print("ESTIMATED TIME ",et-st)
    print("main fn running...")
    '''


    main()

MAIN_FUNCTION()












